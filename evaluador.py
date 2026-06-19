"""
Lógica de evaluación BEP/EIR - Posgrado BIM FADU
Usa Google Gemini (API gratuita).
"""

import re
import json
import unicodedata
import io
import openpyxl
import pdfplumber
from docx import Document
import google.generativeai as genai

MODELO    = "gemini-1.5-flash"   # gratis, rápido, muy capaz
MAX_CHARS = 60_000


# ── Criterios EIR ─────────────────────────────────────────────────────────────

EIR_CRITERIA = [
    ("Historial",                                          "El alumno tiene historial de entregas previas en la cursada"),
    ("Plantilla",                                          "Utilizó la plantilla provista por la cátedra y respetó su estructura"),
    ("Formato",                                            "Formato adecuado: legibilidad, diagramación, PDF correcto"),
    ("Introducción / Objetivo del doc.",                   "Incluye introducción clara con el objetivo del documento EIR"),
    ("Información del proyecto - Nombre",                  "Nombre del proyecto completo y correcto"),
    ("Información del proyecto - Código",                  "Código del proyecto indicado"),
    ("Información del proyecto - Ubicación",               "Ubicación del proyecto indicada"),
    ("Información del proyecto - Coordenadas",             "Coordenadas del proyecto indicadas"),
    ("Información del proyecto - Dimensiones",             "Dimensiones del proyecto indicadas"),
    ("Tipo de Proyecto",                                   "Tipo de proyecto correctamente clasificado"),
    ("Tipo de Intervención",                               "Tipo de intervención correctamente clasificado"),
    ("Función del Recurso Físico - Según Actividad",       "Función del recurso físico según actividad principal indicada"),
    ("Etapa del Proyecto - Etapa",                         "Etapa del proyecto correctamente indicada"),
    ("Etapa del Proyecto - Subetapa",                      "Subetapa del proyecto correctamente indicada"),
    ("Requisitos de proyecto - Memoria",                   "Memoria descriptiva del proyecto incluida"),
    ("Requisitos de proyecto - Entorno Significativo",     "Entorno significativo documentado"),
    ("Requisitos de proyecto - Planilla de locales",       "Planilla de locales incluida o referenciada"),
    ("Requisitos de proyecto - Requerimientos especiales", "Requerimientos especiales indicados"),
    ("Requisitos de proyecto - Normativa de aplicación",   "Normativa de aplicación indicada"),
    ("Información responsables - Solicitante",             "Solicitante del proyecto indicado"),
    ("Información responsables - Alumno/s",                "Alumno/s indicados correctamente"),
    ("Información responsables - Otros",                   "Otros responsables indicados si corresponde"),
    ("Documentación de referencia (INPUT)",                "Documentación de referencia (inputs) incluida"),
    ("Manejo de Información - Objetivos del modelo",       "Objetivos del modelo definidos"),
    ("Manejo de Información - Usos",                       "Usos BIM definidos"),
    ("Manejo de Información - Tipos de información",       "Tipos de información definidos"),
    ("Manejo de Información - Niveles de Información",     "Niveles de información (LOI/LOD) definidos"),
    ("Entregables planificados (OUTPUT)",                  "Entregables planificados (outputs) definidos"),
    ("Intercambio - Esquema de reuniones",                 "Esquema de reuniones definido"),
    ("Intercambio - Nomenclatura especif.",                "Nomenclatura específica de intercambio definida"),
    ("Intercambio - CDE (Esquema)",                        "Esquema CDE definido"),
    ("Intercambio - Niveles de acceso",                    "Niveles de acceso definidos"),
    ("Intercambio - Responsabilidad de la información",    "Responsabilidad de la información definida"),
    ("Requisitos Tecnológicos - Software",                 "Software requerido indicado"),
    ("Requisitos Tecnológicos - Hardware",                 "Hardware requerido indicado"),
    ("Cronograma - Documentos",                            "Cronograma de documentos incluido"),
    ("Cronograma - Modelado",                              "Cronograma de modelado incluido"),
    ("Cronograma - Revisión",                              "Cronograma de revisión incluido"),
    ("Otros",                                              "Sección 'Otros' completada o indicada como N/A"),
    ("Presupuesto",                                        "Presupuesto incluido o referenciado"),
    ("Glosario",                                           "Glosario incluido"),
    ("Listado referentes",                                 "Listado de referentes / bibliografía incluido"),
    ("Observaciones para devolución",                      "Sección de observaciones para devolución presente"),
    ("NOTAS",                                              "Sección de notas presente"),
]


# ── Criterios BEP ─────────────────────────────────────────────────────────────

BEP_CRITERIA = [
    ("Historial",                                                "El alumno tiene historial de entregas previas"),
    ("Plantilla",                                                "Utilizó la plantilla BEP provista por la cátedra"),
    ("Formato",                                                  "Formato adecuado: legibilidad, diagramación, PDF correcto"),
    ("Introducción - 1. Objetivo del documento",                 "Objetivo del documento sintetizado correctamente"),
    ("Introducción - 2. BEP etapas y versiones",                 "Etapas y versiones del BEP indicadas"),
    ("Información del proyecto - 3. Información general",        "Información general del proyecto completa"),
    ("Tipos de modelo - 4. OBRA-Arquitectura",                   "Modelo de Arquitectura definido"),
    ("Tipos de modelo - OBRA-Estructura",                        "Modelo de Estructura definido"),
    ("Tipos de modelo - OBRA-Mobiliario / Interiorismo",         "Modelo de Mobiliario/Interiorismo definido"),
    ("Tipos de modelo - Instalaciones-MEP",                      "Modelo MEP definido"),
    ("Tipos de modelo - SITIO-Entorno / Urbanismo",              "Modelo de Sitio/Entorno definido"),
    ("Tipos de modelo - Coordinación",                           "Modelo de Coordinación definido"),
    ("Tipos de modelo - OBRA-Análisis Energético",               "Modelo de Análisis Energético definido"),
    ("Tipos de modelo - Ingenierías adicionales",                "Modelos de ingenierías adicionales definidos"),
    ("Tipos de modelo - Otros",                                  "Otros tipos de modelo definidos o N/A"),
    ("Manejo de Información - 5. Objetivos del modelado",        "Objetivos del modelado definidos"),
    ("Manejo de Información - Usos",                             "Usos BIM definidos"),
    ("Manejo de Información - Niveles de Información",           "Niveles de información definidos"),
    ("Manejo de Información - Tipos de información",             "Tipos de información definidos"),
    ("Manejo de Información - Niveles geométrico",               "Niveles de desarrollo geométrico definidos"),
    ("Manejo de Información - Niveles no geométrico",            "Niveles de desarrollo no geométrico definidos"),
    ("Estrategia de Modelado - 6. Estrategia BIM",               "Estrategia BIM definida"),
    ("Particularidades - 7. Link CMD",                           "Link al CMD incluido"),
    ("Particularidades - Criterios específicos o desvíos CMD",   "Criterios específicos o desvíos del CMD indicados"),
    ("Particularidades - Flujo de progresión de modelado",       "Flujo de progresión de modelado de cada tipo de modelo"),
    ("Particularidades - Condicionantes o Desafíos",             "Condicionantes o desafíos del modelado indicados"),
    ("Fases de Proyecto - Programación de obra",                 "Programación de obra incluida"),
    ("Fases de Proyecto - Certificación",                        "Certificación incluida"),
    ("Fases de Proyecto - 8. Bitácora",                          "Bitácora incluida"),
    ("Fases de Proyecto - 9. INPUT-BEP",                         "Información de Referencia (INPUT-BEP) incluida"),
    ("Fases de Proyecto - 10. OUTPUT",                           "Productos o Entregables (OUTPUT) definidos"),
    ("Estrategia de revisión - 11. Coordinación y Control",      "Procesos de coordinación y control definidos"),
    ("Estrategia de revisión - Manual",                          "Revisión manual definida"),
    ("Estrategia de revisión - Automático",                      "Revisión automática definida"),
    ("Estrategia de revisión - Dynamo",                          "Uso de Dynamo definido o N/A"),
    ("Estrategia de revisión - Softwares",                       "Softwares de revisión indicados"),
    ("Cronograma - 12. Documentos",                              "Cronograma de documentos incluido"),
    ("Cronograma - Modelado",                                    "Cronograma de modelado incluido"),
    ("Cronograma - Revisión",                                    "Cronograma de revisión incluido"),
    ("Estándares - 13. Proyecto y construcción",                 "Estándares de proyecto y construcción indicados"),
    ("Estándares - Criterios Modelado",                          "Criterios de modelado indicados"),
    ("Estándares - Clasificación y Denominación",                "Clasificación y denominación indicados"),
    ("Estándares - Manejo de información",                       "Manejo de información indicado"),
    ("Estándares - Otras",                                       "Otras normas indicadas o N/A"),
    ("14. Plantillas",                                           "Plantillas de vista/otros incluidas"),
    ("15. Componentes",                                          "Componentes/familias definidos"),
    ("16. Tecnología - Software/Hardware",                       "Tecnología (software/hardware) definida"),
    ("17. Modalidad de intercambio",                             "Modalidad de intercambio definida"),
    ("18. CDE - Plataforma de Intercambio",                      "CDE (plataforma de intercambio) definido"),
    ("18. CDE - Estructura de carpetas",                         "Estructura de carpetas del CDE definida"),
    ("18. CDE - Niveles de permisos",                            "Niveles de permisos definidos"),
    ("18. CDE - Responsabilidad de la Info",                     "Responsabilidad de la información definida"),
    ("19. Pautas de comunicación - Reuniones",                   "Reuniones definidas"),
    ("19. Pautas de comunicación - Informales",                  "Comunicaciones informales definidas"),
    ("19. Pautas de comunicación - Formales",                    "Comunicaciones formales definidas"),
    ("19. Pautas de comunicación - Medios",                      "Medios de comunicación definidos"),
    ("20. Roles y responsabilidades",                            "Roles y responsabilidades definidos"),
    ("21. Conclusiones - Proceso",                               "Conclusiones sobre el proceso"),
    ("21. Conclusiones - Producto",                              "Conclusiones sobre el producto"),
    ("21. Conclusiones - Equipos RRHH",                          "Conclusiones sobre equipos y RRHH"),
    ("22. Glosario",                                             "Glosario incluido"),
    ("23. Referentes - Bibliografía",                            "Referentes y bibliografía incluidos"),
    ("Observaciones para devolución",                            "Sección de observaciones para devolución presente"),
    ("NOTAS",                                                    "Sección de notas presente"),
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def normalizar(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


def _strip_nonalpha(s: str) -> str:
    return re.sub(r"[^a-z0-9 ]", "", s)


# ── Lectura de documentos ──────────────────────────────────────────────────────

def _read_pdf(file_like) -> str:
    with pdfplumber.open(file_like) as pdf:
        return "\n".join(p.extract_text() or "" for p in pdf.pages)


def _read_docx(file_like) -> str:
    doc = Document(file_like)
    return "\n".join(p.text for p in doc.paragraphs)


def _read_xlsx(file_like) -> str:
    wb = openpyxl.load_workbook(file_like, data_only=True)
    lines = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        lines.append(f"[Hoja: {sn}]")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) for c in row if c is not None]
            if vals:
                lines.append(" | ".join(vals))
    return "\n".join(lines)


def extract_text(uploaded_file) -> str:
    """Extrae texto de un archivo subido por Streamlit."""
    name = uploaded_file.name.lower()
    try:
        if name.endswith(".pdf"):
            return _read_pdf(uploaded_file)[:MAX_CHARS]
        elif name.endswith(".docx"):
            return _read_docx(uploaded_file)[:MAX_CHARS]
        elif name.endswith((".xlsx", ".xlsm", ".xls")):
            return _read_xlsx(uploaded_file)[:MAX_CHARS]
    except Exception as e:
        return f"[Error leyendo {uploaded_file.name}: {e}]"
    return ""


def clasificar_archivos(uploaded_files: list) -> tuple[list, list]:
    """Separa archivos en (eir_files, bep_files) según el nombre."""
    eir, bep = [], []
    for f in uploaded_files:
        fl = f.name.lower()
        if "eir" in fl:
            eir.append(f)
        elif "bep" in fl:
            bep.append(f)
        else:
            # Si no tiene etiqueta clara, lo incluimos en ambos
            eir.append(f)
            bep.append(f)
    return eir, bep


# ── Evaluación con Gemini ──────────────────────────────────────────────────────

def _build_prompt(doc_type: str, doc_text: str, criteria: list) -> str:
    criteria_json = json.dumps(
        [{"id": i, "criterio": c[0], "descripción": c[1]} for i, c in enumerate(criteria)],
        ensure_ascii=False, indent=2
    )
    texto = doc_text.strip() or "[Documento vacío o no provisto]"
    return f"""Sos docente correctora del posgrado BIM FADU.
Evaluás el siguiente documento {doc_type} de un alumno según estos {len(criteria)} criterios:

{criteria_json}

DOCUMENTO:
{texto}

Respondé SOLO con JSON válido (sin texto extra) con esta estructura:
{{
  "resultados": [
    {{"id": 0, "criterio": "nombre", "estado": "SI|NO|REV|INC|N/C", "obs": "observación breve"}}
  ]
}}

Estados: SI=correcto, NO=ausente/incorrecto, REV=requiere revisión, INC=incompleto, N/C=no corresponde.
Obs vacío si estado es SI sin problemas. Observaciones concretas y útiles para el alumno."""


def evaluar_documento(api_key: str, doc_type: str, doc_text: str, criteria: list) -> list:
    if not doc_text.strip():
        return [{"id": i, "criterio": c[0], "estado": "N/C", "obs": "Documento no provisto"}
                for i, c in enumerate(criteria)]

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(MODELO)
    response = model.generate_content(_build_prompt(doc_type, doc_text, criteria))
    raw = response.text.strip()

    m = re.search(r'\{.*\}', raw, re.DOTALL)
    if not m:
        raise ValueError(f"Respuesta inesperada: {raw[:300]}")
    return json.loads(m.group())["resultados"]


# ── Planilla Excel ─────────────────────────────────────────────────────────────

def _detect_name_row(ws) -> int:
    SKIP = {"EIR", "BEP", "REC", "ESTADO", "OBS.", "N°", "APELLIDO", "NOMBRE"}
    best_row, best_count = 1, 0
    for row_idx in range(1, 5):
        count = sum(
            1 for cell in ws[row_idx]
            if cell.value and isinstance(cell.value, str)
            and cell.value.upper().strip() not in SKIP
        )
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def build_column_map(ws) -> tuple[dict, int]:
    """Devuelve ({apellido_norm: (col_estado, col_obs)}, name_row)."""
    SKIP = {"EIR", "BEP", "REC", "ESTADO", "OBS."}
    name_row = _detect_name_row(ws)
    col_map = {}
    cells = [
        cell for cell in ws[name_row]
        if cell.value and isinstance(cell.value, str) and cell.value.upper() not in SKIP
    ]
    i = 0
    while i < len(cells):
        cell = cells[i]
        col = cell.column
        key = normalizar(cell.value)
        col_map[key] = (col, col + 1)
        i += 1
        if i < len(cells) and cells[i].column == col + 1:
            i += 1
    return col_map, name_row


def find_student_cols(col_map: dict, apellido: str):
    key = normalizar(apellido)
    if key in col_map:
        return col_map[key]
    key_clean = _strip_nonalpha(key).replace(" ", "")
    for k, v in col_map.items():
        if _strip_nonalpha(k).replace(" ", "") == key_clean:
            return v
    if len(key) >= 4:
        for k, v in col_map.items():
            if key in k or k in key:
                return v
    return None, None


def get_student_list(planilla_bytes: bytes) -> list[dict]:
    """Lee la planilla desde bytes y devuelve lista de alumnos."""
    wb = openpyxl.load_workbook(io.BytesIO(planilla_bytes), data_only=True)
    ws_eir = wb["EIR"]
    ws_bep = wb["BEP"]
    col_map_eir, name_row_eir = build_column_map(ws_eir)
    col_map_bep, _            = build_column_map(ws_bep)

    SKIP = {"EIR", "BEP", "REC", "ESTADO", "OBS."}
    students = []
    cells = [
        cell for cell in ws_eir[name_row_eir]
        if cell.value and isinstance(cell.value, str) and cell.value.upper() not in SKIP
    ]
    i = 0
    while i < len(cells):
        ap_cell = cells[i]
        apellido = ap_cell.value
        nombre = ""
        if i + 1 < len(cells) and cells[i + 1].column == ap_cell.column + 1:
            nombre = cells[i + 1].value or ""
            i += 2
        else:
            i += 1
        if normalizar(apellido) in {"eir", "bep", "rec"}:
            continue
        students.append({
            "apellido": apellido,
            "nombre":   nombre,
            "display":  f"{apellido}, {nombre}".strip(", "),
            "col_eir":  col_map_eir.get(normalizar(apellido), (None, None)),
            "col_bep":  col_map_bep.get(normalizar(apellido), (None, None)),
        })
    return students


def _row_for_criterion(ws, criterio_name: str):
    base = normalizar(criterio_name.split(" - ")[0])
    for row in ws.iter_rows(min_col=2, max_col=3):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if base in normalizar(cell.value):
                    return cell.row
    return None


def escribir_resultados(planilla_bytes: bytes, apellido: str,
                        eir_results: list, bep_results: list) -> bytes:
    """Escribe los resultados en la planilla y devuelve los bytes actualizados."""
    wb = openpyxl.load_workbook(io.BytesIO(planilla_bytes))
    ws_eir = wb["EIR"]
    ws_bep = wb["BEP"]

    col_map_eir, _ = build_column_map(ws_eir)
    col_map_bep, _ = build_column_map(ws_bep)

    for ws, results, criteria, col_map in [
        (ws_eir, eir_results, EIR_CRITERIA, col_map_eir),
        (ws_bep, bep_results, BEP_CRITERIA, col_map_bep),
    ]:
        col_e, col_o = find_student_cols(col_map, apellido)
        if not col_e:
            continue
        for r in results:
            idx = r.get("id", -1)
            if not (0 <= idx < len(criteria)):
                continue
            row = _row_for_criterion(ws, criteria[idx][0])
            if row:
                ws.cell(row=row, column=col_e).value = r.get("estado", "")
                if r.get("obs"):
                    ws.cell(row=row, column=col_o).value = r["obs"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
